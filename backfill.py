#!/usr/bin/env python3
"""ONE-OFF history backfill for the ATM/POS/Card dashboard — ERA-AWARE (handles old + new RBI layouts)."""
import os, io, re, time, datetime
import build_dashboard as B  # read_embedded, splice, month_sort_key, DOC_RE, LIST_URL, UA, HTML_PATH, SCHEMA

START = int(os.environ.get("START_ATMID", "40"))
END   = int(os.environ.get("END_ATMID", "190"))
START_YEAR = int(os.environ.get("START_YEAR", "2017"))

MAB={'january':'Jan','february':'Feb','march':'Mar','april':'Apr','may':'May','june':'Jun','july':'Jul',
     'august':'Aug','september':'Sep','october':'Oct','november':'Nov','december':'Dec'}
GROUPS={"public sector banks":"Public Sector","private sector banks":"Private Sector",
        "foreign banks":"Foreign","payment banks":"Payments","payments banks":"Payments",
        "small finance banks":"Small Finance"}

def _num(x):
    if x is None: return 0
    if isinstance(x,(int,float)): return x
    s=str(x).strip().replace(",","")
    if s=="" or s=="-": return 0
    try: return float(s)
    except Exception: return 0

def _parse_new(rows):
    banks={}; total=None; cur=None; groups={}
    for r in rows:
        c1=r[1] if len(r)>1 else None; c2=r[2] if len(r)>2 else None
        if c1 is None and c2 is None: continue
        s1=str(c1).strip() if c1 is not None else ""
        if s1.lower()=="total" and (c2 is None or str(c2).strip()==""):
            total=[round(_num(r[c]),3) if c<len(r) else 0 for c in range(3,29)]; continue
        if c2 is None or str(c2).strip()=="":
            g=GROUPS.get(s1.lower())
            if g: cur=g
            continue
        name=str(c2).strip()
        if not name or name.lower()=="bank name": continue
        banks[name]=[round(_num(r[c]),3) if c<len(r) else 0 for c in range(3,29)]
        groups[name]=cur or "Other"
    return banks, total, groups

def _old_row(r):
    def g(i): return _num(r[i]) if i<len(r) else 0
    return [round(x,3) for x in [
        g(2), g(3), g(4)+g(5), g(6), g(7), 0, g(8), g(13),
        g(10), g(12)*100, 0,0, 0,0, g(9), g(11)*100,
        g(15), g(17)*100, 0,0, 0,0, g(14), g(16)*100, 0,0]]

def _parse_old(rows):
    banks={}; total=None; cur=None; groups={}
    for r in rows:
        c1=r[1] if len(r)>1 else None; c2=r[2] if len(r)>2 else None
        if c1 is None: continue
        s1=str(c1).strip()
        if s1.lower()=="total":
            total=_old_row(r); continue
        if c2 is None or str(c2).strip()=="":
            g=GROUPS.get(s1.lower())
            if g: cur=g
            continue
        if not isinstance(c2,(int,float)): continue
        if s1.lower()=="bank name": continue
        banks[s1]=_old_row(r); groups[s1]=cur or "Other"
    return banks, total, groups

def detect_format(rows):
    txt=" ".join(str(c) for r in rows[:7] for c in (r or []) if c is not None)
    if "UPI QR" in txt: return "new"
    if ("Rupees Lakh" in txt) or ("On-line" in txt and "Off-line" in txt): return "old"
    if "Sr. No" in txt or "Sr.No" in txt: return "new"
    return None

def parse_any(b):
    import openpyxl
    wb=openpyxl.load_workbook(io.BytesIO(b), read_only=True, data_only=True)
    sn=wb.sheetnames[0]
    m=re.search(r"([A-Za-z]+)\s+(\d{4})", sn)
    if not m: return None
    mon=MAB.get(m.group(1).lower())
    if not mon: return None
    mk=f"{mon} {m.group(2)}"
    rows=list(wb[sn].iter_rows(values_only=True))
    fmt=detect_format(rows)
    if fmt=="new": banks,total,groups=_parse_new(rows)
    elif fmt=="old": banks,total,groups=_parse_old(rows)
    else: return ("UNRECOGNISED", mk)
    if not banks: return ("UNRECOGNISED", mk)
    return mk, {"rows":banks,"total":total}, groups

def main():
    from playwright.sync_api import sync_playwright
    html=open(B.HTML_PATH, encoding="utf-8").read()
    cur=B.read_embedded(html); months=dict(cur.get("months",{})); groups=dict(cur.get("groups",{}))
    before_n=len(months)
    links=set()
    with sync_playwright() as p:
        br=p.chromium.launch(args=["--no-sandbox"]); ctx=br.new_context(user_agent=B.UA); pg=ctx.new_page()
        pg.goto(B.LIST_URL, wait_until="domcontentloaded", timeout=90000); pg.wait_for_timeout(2500)
        links |= set(B.DOC_RE.findall(pg.content()))
        for N in range(END, START-1, -1):
            try:
                pg.goto(B.LIST_URL+"?atmid=%d"%N, wait_until="domcontentloaded", timeout=45000)
                links |= set(B.DOC_RE.findall(pg.content()))
            except Exception as e:
                print("  atmid", N, "skipped:", e)
        print("Collected %d unique monthly file link(s)." % len(links))
        files={}
        for u in sorted(links):
            try:
                r=ctx.request.get(u, headers={"Referer":B.LIST_URL}, timeout=90000); b=r.body()
                if b[:2]==b"PK": files[u]=b
            except Exception as e:
                print("  download failed:", u[-34:], e)
        br.close()
    added=0; skipped=[]
    for u,b in files.items():
        try:
            res=parse_any(b)
            if not res: continue
            if res[0]=="UNRECOGNISED":
                skipped.append(res[1]); continue
            mk,md,g=res
            if int(mk.split()[1])<START_YEAR: continue
            months[mk]=md; groups.update(g); added+=1
        except Exception as e:
            print("  parse error:", e)
    order=sorted(months, key=B.month_sort_key)
    payload={"version":1,"exportedAt":datetime.datetime.utcnow().isoformat()+"Z",
             "source":"RBI Bankwise ATM/POS/Card Statistics","schema":cur.get("schema",B.SCHEMA),
             "valueUnit":"Rs'000","groups":groups,"monthOrder":order,"months":months}
    html=B.splice(html, payload, int(time.time()*1000), order[-1] if order else "")
    open(B.HTML_PATH,"w",encoding="utf-8").write(html)
    print("\n==== BACKFILL SUMMARY ====")
    print("Downloaded files       :", len(files))
    print("Months ingested        :", added)
    if skipped:
        sk=sorted(set(x for x in skipped if x))
        print("Skipped (unrecognised) :", len(sk), "->", ", ".join(sk[:24]) + (" …" if len(sk)>24 else ""))
    print("Dashboard now has %d month(s): %s -> %s" % (len(order), order[0] if order else "-", order[-1] if order else "-"))
    print("Was %d before backfill." % before_n)

if __name__ == "__main__":
    main()
