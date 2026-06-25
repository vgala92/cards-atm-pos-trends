#!/usr/bin/env python3
"""
Auto-update the RBI Bankwise ATM/POS/Card dashboard.

Unlike the daily NFS file (one stable URL), this report has ONE Excel per month and the
download links carry a random hash, so the script must:
  1. Scrape the listing page (ATMView.aspx) for the monthly .XLSX links.
  2. Download each listed month's file (handles RBI's bot wall via a headless browser).
  3. Parse the NFS-style bank rows, MERGE into the data already embedded in index.html
     (so older months are retained and "Revised" months are refreshed).
  4. Splice the merged data + a new version stamp + "published" label into index.html.

If nothing changed, index.html is left untouched and the script exits 0.
The dashboard (index.html) IS the template: only EMBEDDED_DATA / EMBEDDED_STAMP /
EMBEDDED_PUBLISHED are rewritten.
"""

import os, re, sys, json, time, io, datetime

LIST_URL = "https://rbi.org.in/Scripts/ATMView.aspx"
DOC_RE   = re.compile(r"https://rbidocs\.rbi\.org\.in/rdocs/ATM/DOCs/ATM[^\"'<>]+?\.XLSX", re.I)
HTML_PATH = os.path.join(os.path.dirname(__file__), "index.html")
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36")

SCHEMA = ["atm_on","atm_off","pos","micro_atm","bharat_qr","upi_qr","cc_out","dc_out",
          "cc_pos_v","cc_pos_val","cc_online_v","cc_online_val","cc_other_v","cc_other_val","cc_atm_v","cc_atm_val",
          "dc_pos_v","dc_pos_val","dc_online_v","dc_online_val","dc_other_v","dc_other_val",
          "dc_atm_v","dc_atm_val","dc_poscw_v","dc_poscw_val"]
MAB={'january':'Jan','february':'Feb','march':'Mar','april':'Apr','may':'May','june':'Jun','july':'Jul',
     'august':'Aug','september':'Sep','october':'Oct','november':'Nov','december':'Dec'}
GROUPS={"public sector banks":"Public Sector","private sector banks":"Private Sector",
        "foreign banks":"Foreign","payment banks":"Payments","payments banks":"Payments",
        "small finance banks":"Small Finance"}


# ---------------------------------------------------------------- fetch
def _is_xlsx(b): return bool(b) and b[:2] == b'PK'

def get_listing_html():
    """The .aspx listing page is not behind the bot wall — plain request works."""
    import requests
    r = requests.get(LIST_URL, headers={"User-Agent": UA}, timeout=60)
    return r.text

def fetch_via_playwright(urls):
    """Download each file inside a real browser context that clears Incapsula."""
    out = {}
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        browser = p.chromium.launch(args=["--no-sandbox"])
        ctx = browser.new_context(user_agent=UA, accept_downloads=True)
        page = ctx.new_page()
        page.goto(LIST_URL, wait_until="networkidle", timeout=90000)
        page.wait_for_timeout(2500)
        for u in urls:
            try:
                resp = ctx.request.get(u, headers={"Referer": LIST_URL}, timeout=90000)
                body = resp.body()
                if _is_xlsx(body): out[u] = body
                else: print("  not xlsx:", u[-40:], "head=", body[:8])
            except Exception as e:
                print("  download failed:", u[-40:], e)
        browser.close()
    return out

def fetch_files(urls):
    out = {}
    # Strategy 1 — plain requests (works if not challenged)
    try:
        import requests
        s = requests.Session(); s.headers.update({"User-Agent": UA, "Referer": LIST_URL})
        try: s.get(LIST_URL, timeout=60)  # prime cookies
        except Exception: pass
        for u in urls:
            try:
                r = s.get(u, timeout=90)
                if _is_xlsx(r.content): out[u] = r.content
            except Exception: pass
        if len(out) == len(urls): print("Fetched all via requests"); return out
    except Exception as e:
        print("requests path:", e)
    # Strategy 2 — Playwright for the rest
    missing = [u for u in urls if u not in out]
    if missing:
        print("Fetching %d file(s) via headless browser…" % len(missing))
        out.update(fetch_via_playwright(missing))
    return out


# ---------------------------------------------------------------- parse
def _num(x):
    if x is None: return 0
    if isinstance(x,(int,float)): return x
    s=str(x).strip().replace(",","")
    if s=="" or s=="-": return 0
    try: return float(s)
    except Exception: return 0

def parse_bytes(b):
    import openpyxl
    wb=openpyxl.load_workbook(io.BytesIO(b), read_only=True, data_only=True)
    sn=wb.sheetnames[0]
    m=re.search(r"([A-Za-z]+)\s+(\d{4})", sn)
    if not m: return None
    mon=MAB.get(m.group(1).lower())
    if not mon: return None
    mk=f"{mon} {m.group(2)}"
    rows=list(wb[sn].iter_rows(values_only=True))
    banks={}; total=None; cur=None; groups={}
    for r in rows:
        c1=r[1] if len(r)>1 else None
        c2=r[2] if len(r)>2 else None
        if c1 is None and c2 is None: continue
        s1=str(c1).strip() if c1 is not None else ""
        if s1.lower()=="total" and (c2 is None or str(c2).strip()==""):
            total=[round(_num(r[c]),3) if c<len(r) else 0 for c in range(3,29)]; continue
        if c2 is None or str(c2).strip()=="":
            g=GROUPS.get(s1.lower())
            if g: cur=g
            continue
        name=str(c2).strip()
        if not name or name.lower()=="bank name": continue
        banks[name]=[round(_num(r[c]),3) if c<len(r) else 0 for c in range(3,29)]
        groups[name]=cur or "Other"
    if not banks: return None
    return mk, {"rows":banks,"total":total}, groups


# ---------------------------------------------------------------- splice / merge
def month_sort_key(mk):
    mo=["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    a,b=mk.split(); return (int(b), mo.index(a)+1)

def read_embedded(html):
    m=re.search(r"const EMBEDDED_DATA = (\{.*?\});\nconst EMBEDDED_STAMP", html, re.S)
    if not m: return {"months":{},"groups":{},"schema":SCHEMA,"monthOrder":[]}
    try: return json.loads(m.group(1))
    except Exception: return {"months":{},"groups":{},"schema":SCHEMA,"monthOrder":[]}

def splice(html, payload, stamp, published):
    def rb(s, start, end, inner):
        i=s.index(start); j=s.index(end, i+len(start)); return s[:i]+start+inner+s[j:]
    html=rb(html, "const EMBEDDED_DATA = ", ";\nconst EMBEDDED_STAMP", json.dumps(payload, separators=(",",":")))
    html=rb(html, "const EMBEDDED_STAMP = ", ";", str(stamp))
    html=rb(html, 'const EMBEDDED_PUBLISHED = "', '"', published)
    return html


# ---------------------------------------------------------------- main
def main():
    if not os.path.exists(HTML_PATH):
        print("ERROR: index.html missing next to this script."); sys.exit(1)

    try:
        page = get_listing_html()
    except Exception as e:
        print("ERROR: could not load listing page:", e); sys.exit(1)
    urls = sorted(set(DOC_RE.findall(page)))
    if not urls:
        print("ERROR: no .XLSX links found on the ATM listing page."); sys.exit(1)
    print("Found %d monthly file link(s) on the page." % len(urls))

    files = fetch_files(urls)
    if not files:
        print("ERROR: could not download any RBI ATM file (bot protection / network)."); sys.exit(1)

    parsed = {}; groups = {}
    for u, b in files.items():
        try:
            res = parse_bytes(b)
            if res:
                mk, mdata, g = res
                parsed[mk] = mdata; groups.update(g)
        except Exception as e:
            print("  parse failed:", u[-40:], e)
    if not parsed:
        print("ERROR: downloaded files but parsed no months."); sys.exit(1)

    html = open(HTML_PATH, encoding="utf-8").read()
    cur = read_embedded(html)
    months = dict(cur.get("months", {}))
    allgroups = dict(cur.get("groups", {}))

    # merge: scraped months overwrite (handles revisions); old months retained
    before = json.dumps({"months":months,"monthOrder":sorted(months,key=month_sort_key)}, separators=(",",":"))
    months.update(parsed)
    allgroups.update(groups)
    order = sorted(months.keys(), key=month_sort_key)
    after = json.dumps({"months":months,"monthOrder":order}, separators=(",",":"))

    if before == after:
        print("No new or revised ATM data — index.html left untouched."); return

    published = order[-1]
    payload = {"version":1,"exportedAt":datetime.datetime.utcnow().isoformat()+"Z",
               "source":"RBI Bankwise ATM/POS/Card Statistics","schema":cur.get("schema",SCHEMA),
               "valueUnit":"Rs'000","groups":allgroups,"monthOrder":order,"months":months}
    stamp = int(time.time()*1000)
    html = splice(html, payload, stamp, published)
    open(HTML_PATH, "w", encoding="utf-8").write(html)
    print(f"Updated index.html — {len(order)} months ({order[0]} → {order[-1]}), stamp {stamp}")


if __name__ == "__main__":
    main()
